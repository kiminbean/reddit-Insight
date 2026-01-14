"""Excel Generator 테스트.

ExcelGenerator 클래스의 Excel 생성 기능을 테스트한다.
"""

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pytest

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# TEST FIXTURES
# =============================================================================


@dataclass
class BusinessItem:
    """테스트용 BusinessItem."""

    rank: int
    title: str
    category: str
    opportunity_score: float
    market_potential: str
    risk_level: str
    description: str
    target_audience: str
    key_features: list[str]
    competitive_advantage: str
    next_steps: list[str]
    evidence: list[str]


@dataclass
class MockReportData:
    """테스트용 ReportData mock."""

    subreddit: str
    generated_at: datetime
    analysis_period: str
    total_posts_analyzed: int
    total_keywords: int
    total_insights: int
    executive_summary: str
    market_overview: dict[str, Any]
    business_items: list[BusinessItem]
    trend_analysis: dict[str, Any]
    demand_analysis: dict[str, Any]
    competition_analysis: dict[str, Any]
    recommendations: list[str]
    risk_factors: list[str]
    conclusion: str


@pytest.fixture
def sample_report_data() -> MockReportData:
    """샘플 리포트 데이터를 생성한다."""
    return MockReportData(
        subreddit="python",
        generated_at=datetime.now(UTC),
        analysis_period="Last 7 days",
        total_posts_analyzed=150,
        total_keywords=50,
        total_insights=10,
        executive_summary=(
            "r/python 커뮤니티의 150개 게시물 분석 결과, "
            "10개의 인사이트와 3개의 비즈니스 기회를 도출했습니다.\n\n"
            "**최우선 기회**: AI 기반 코드 분석 도구"
        ),
        market_overview={
            "community_size": "r/python",
            "activity_level": "활발",
            "data_quality": "충분",
            "market_maturity": "성장 시장 - 일부 솔루션 존재, 기회 있음",
            "key_topics": ["python", "ai", "machine learning", "automation"],
        },
        business_items=[
            BusinessItem(
                rank=1,
                title="AI 기반 코드 분석 도구",
                category="신규 시장 진입",
                opportunity_score=85.5,
                market_potential="높음",
                risk_level="낮음",
                description="개발자를 위한 AI 기반 코드 품질 분석 및 개선 도구",
                target_audience="소프트웨어 개발자 및 기술 전문가",
                key_features=["AI 기반 자동화", "실시간 처리", "외부 서비스 연동"],
                competitive_advantage="시장 내 미개척 영역 선점으로 인한 선발자 우위",
                next_steps=[
                    "MVP 개발 및 초기 사용자 테스트 진행",
                    "핵심 타겟 고객 인터뷰 실시",
                ],
                evidence=[
                    "I wish there was a tool that could automatically review my code",
                    "We need better AI integration in our development workflow",
                ],
            ),
            BusinessItem(
                rank=2,
                title="파이썬 학습 플랫폼",
                category="트렌드 기반 서비스",
                opportunity_score=72.0,
                market_potential="중간",
                risk_level="중간",
                description="초보자를 위한 인터랙티브 파이썬 학습 플랫폼",
                target_audience="학생 및 교육 관련 사용자",
                key_features=["실시간 처리", "협업 기능"],
                competitive_advantage="트렌드 선도를 통한 시장 리더십 확보",
                next_steps=[
                    "시장 조사 및 수요 검증 심화",
                    "프로토타입 개발 및 피드백 수집",
                ],
                evidence=[
                    "Learning Python is hard without proper guidance",
                ],
            ),
        ],
        trend_analysis={
            "total_keywords": 50,
            "top_keywords": [
                {"keyword": "python", "score": 0.95},
                {"keyword": "ai", "score": 0.88},
                {"keyword": "machine learning", "score": 0.82},
                {"keyword": "automation", "score": 0.75},
                {"keyword": "data science", "score": 0.70},
            ],
            "rising_topics": [
                {"topic": "LLM", "trend": "rising"},
                {"topic": "RAG", "trend": "rising"},
            ],
            "trend_summary": "주요 키워드: python, ai, machine learning.",
        },
        demand_analysis={
            "total_demands": 25,
            "by_category": {"feature_request": 10, "pain_point": 8, "search_query": 7},
            "top_opportunities": [
                {
                    "representative": "Better debugging tools for Python",
                    "priority_score": 75,
                    "business_potential": "high",
                },
                {
                    "representative": "AI-powered code completion",
                    "priority_score": 70,
                    "business_potential": "medium",
                },
            ],
            "demand_summary": "총 25개의 수요 신호가 감지되었습니다.",
        },
        competition_analysis={
            "entities_mentioned": [],
            "sentiment_distribution": {},
            "key_complaints": [],
            "competition_summary": "경쟁 분석 데이터가 충분하지 않습니다.",
        },
        recommendations=[
            "1. **AI 기반 코드 분석 도구** 기회를 우선 검토하세요.",
            "2. 사용자 수요가 높은 기능에 집중하여 MVP를 설계하세요.",
            "3. 타겟 커뮤니티에서 직접 피드백을 수집하세요.",
        ],
        risk_factors=[
            "데이터 샘플 크기가 작아 분석 결과의 신뢰도가 제한될 수 있음",
            "Reddit 커뮤니티 의견이 전체 시장을 대표하지 않을 수 있음",
        ],
        conclusion=(
            "본 분석을 통해 2개의 비즈니스 기회를 도출했습니다. "
            "제시된 추천 사항을 참고하여 시장 검증을 진행하시기 바랍니다."
        ),
    )


@pytest.fixture
def empty_report_data() -> MockReportData:
    """빈 리포트 데이터를 생성한다."""
    return MockReportData(
        subreddit="empty_sub",
        generated_at=datetime.now(UTC),
        analysis_period="Last 7 days",
        total_posts_analyzed=10,
        total_keywords=0,
        total_insights=0,
        executive_summary="분석 데이터가 충분하지 않습니다.",
        market_overview={
            "community_size": "r/empty_sub",
            "activity_level": "낮음",
            "data_quality": "추가 필요",
        },
        business_items=[],
        trend_analysis={
            "total_keywords": 0,
            "top_keywords": [],
            "rising_topics": [],
            "trend_summary": "트렌드 데이터가 없습니다.",
        },
        demand_analysis={
            "total_demands": 0,
            "by_category": {},
            "top_opportunities": [],
            "demand_summary": "수요 분석 데이터가 없습니다.",
        },
        competition_analysis={
            "competition_summary": "경쟁 분석 데이터가 없습니다.",
        },
        recommendations=[],
        risk_factors=[],
        conclusion="데이터가 충분하지 않습니다.",
    )


# =============================================================================
# IMPORT TESTS
# =============================================================================


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestExcelGeneratorImport:
    """ExcelGenerator import 테스트."""

    def test_import_excel_generator(self) -> None:
        """ExcelGenerator를 import할 수 있다."""
        from reddit_insight.reports.excel_generator import ExcelGenerator

        assert ExcelGenerator is not None

    def test_import_utility_function(self) -> None:
        """generate_excel_from_report 유틸리티 함수를 import할 수 있다."""
        from reddit_insight.reports.excel_generator import generate_excel_from_report

        assert generate_excel_from_report is not None


# =============================================================================
# EXCEL GENERATION TESTS
# =============================================================================


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestExcelGenerator:
    """ExcelGenerator 테스트."""

    @pytest.fixture
    def generator(self):
        """ExcelGenerator 인스턴스를 생성한다."""
        from reddit_insight.reports.excel_generator import ExcelGenerator

        return ExcelGenerator()

    def test_generator_repr(self, generator) -> None:
        """ExcelGenerator repr이 올바르게 반환된다."""
        assert repr(generator) == "ExcelGenerator()"

    def test_generate_returns_bytes(self, generator, sample_report_data) -> None:
        """generate가 바이트를 반환한다."""
        result = generator.generate(sample_report_data)

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_creates_valid_excel(self, generator, sample_report_data) -> None:
        """generate가 유효한 Excel 파일을 생성한다."""
        result = generator.generate(sample_report_data)

        # Load workbook to verify it's valid
        wb = load_workbook(BytesIO(result))
        assert wb is not None
        assert len(wb.sheetnames) > 0

    def test_generate_creates_expected_sheets(
        self, generator, sample_report_data
    ) -> None:
        """생성된 Excel이 예상된 시트들을 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))

        expected_sheets = ["Summary", "Opportunities", "Keywords", "Trends", "Demands"]
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Sheet '{sheet_name}' not found"

    def test_generate_with_empty_data(self, generator, empty_report_data) -> None:
        """빈 데이터로도 Excel을 생성할 수 있다."""
        result = generator.generate(empty_report_data)

        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        assert "Summary" in wb.sheetnames


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestSummarySheet:
    """Summary 시트 테스트."""

    @pytest.fixture
    def generator(self):
        """ExcelGenerator 인스턴스를 생성한다."""
        from reddit_insight.reports.excel_generator import ExcelGenerator

        return ExcelGenerator()

    def test_summary_contains_title(self, generator, sample_report_data) -> None:
        """Summary 시트가 제목을 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]

        title_cell = ws["A1"].value
        assert sample_report_data.subreddit in title_cell

    def test_summary_contains_metadata(self, generator, sample_report_data) -> None:
        """Summary 시트가 메타데이터를 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]

        # Find metadata in cells
        all_values = [
            str(cell.value) if cell.value else ""
            for row in ws.iter_rows(min_row=1, max_row=20)
            for cell in row
        ]
        all_text = " ".join(all_values)

        assert "Generated At" in all_text or str(sample_report_data.total_posts_analyzed) in all_text

    def test_summary_contains_recommendations(
        self, generator, sample_report_data
    ) -> None:
        """Summary 시트가 추천 사항을 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]

        # Find recommendations in cells
        all_values = [
            str(cell.value) if cell.value else ""
            for row in ws.iter_rows()
            for cell in row
        ]
        all_text = " ".join(all_values)

        assert "Recommendations" in all_text


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestOpportunitiesSheet:
    """Opportunities 시트 테스트."""

    @pytest.fixture
    def generator(self):
        """ExcelGenerator 인스턴스를 생성한다."""
        from reddit_insight.reports.excel_generator import ExcelGenerator

        return ExcelGenerator()

    def test_opportunities_has_headers(self, generator, sample_report_data) -> None:
        """Opportunities 시트가 헤더를 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Opportunities"]

        headers = [cell.value for cell in ws[1]]
        assert "Rank" in headers
        assert "Title" in headers
        assert "Score" in headers

    def test_opportunities_has_data(self, generator, sample_report_data) -> None:
        """Opportunities 시트가 데이터를 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Opportunities"]

        # Check row 2 (first data row)
        assert ws.cell(row=2, column=1).value == 1  # Rank
        assert sample_report_data.business_items[0].title in str(
            ws.cell(row=2, column=2).value
        )


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestKeywordsSheet:
    """Keywords 시트 테스트."""

    @pytest.fixture
    def generator(self):
        """ExcelGenerator 인스턴스를 생성한다."""
        from reddit_insight.reports.excel_generator import ExcelGenerator

        return ExcelGenerator()

    def test_keywords_has_headers(self, generator, sample_report_data) -> None:
        """Keywords 시트가 헤더를 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Keywords"]

        headers = [cell.value for cell in ws[1]]
        assert "Rank" in headers
        assert "Keyword" in headers
        assert "Score" in headers

    def test_keywords_has_data(self, generator, sample_report_data) -> None:
        """Keywords 시트가 키워드 데이터를 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Keywords"]

        # Check first keyword
        first_keyword = sample_report_data.trend_analysis["top_keywords"][0]
        assert ws.cell(row=2, column=2).value == first_keyword["keyword"]


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestDemandsSheet:
    """Demands 시트 테스트."""

    @pytest.fixture
    def generator(self):
        """ExcelGenerator 인스턴스를 생성한다."""
        from reddit_insight.reports.excel_generator import ExcelGenerator

        return ExcelGenerator()

    def test_demands_has_summary(self, generator, sample_report_data) -> None:
        """Demands 시트가 요약을 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Demands"]

        assert ws["A1"].value == "Demand Summary"

    def test_demands_has_category_distribution(
        self, generator, sample_report_data
    ) -> None:
        """Demands 시트가 카테고리 분포를 포함한다."""
        result = generator.generate(sample_report_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Demands"]

        # Find category distribution in cells
        all_values = [
            str(cell.value) if cell.value else ""
            for row in ws.iter_rows()
            for cell in row
        ]
        all_text = " ".join(all_values)

        assert "Category" in all_text or "feature_request" in all_text


# =============================================================================
# UTILITY FUNCTION TESTS
# =============================================================================


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestGenerateExcelFromReport:
    """generate_excel_from_report 유틸리티 함수 테스트."""

    def test_utility_function_works(self, sample_report_data) -> None:
        """유틸리티 함수가 Excel을 생성한다."""
        from reddit_insight.reports.excel_generator import generate_excel_from_report

        result = generate_excel_from_report(sample_report_data)

        assert isinstance(result, bytes)

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) > 0
