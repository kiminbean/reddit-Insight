"""
Excel Report Generator Module.

Excel 형식의 리포트 생성기를 제공한다.
openpyxl을 사용하여 데이터를 Excel 워크북으로 변환한다.

Example:
    >>> from reddit_insight.reports import ExcelGenerator
    >>> from reddit_insight.dashboard.services.report_service import ReportData
    >>> generator = ExcelGenerator()
    >>> excel_bytes = generator.generate(report_data)
    >>> with open("report.xlsx", "wb") as f:
    ...     f.write(excel_bytes)
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import TYPE_CHECKING, Any

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

if TYPE_CHECKING:
    from reddit_insight.dashboard.services.report_service import ReportData


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================


class ExcelStyles:
    """Excel 스타일 정의."""

    # Fonts
    HEADER_FONT = Font(bold=True, size=14, color="FFFFFF") if OPENPYXL_AVAILABLE else None
    SUBHEADER_FONT = Font(bold=True, size=12) if OPENPYXL_AVAILABLE else None
    TABLE_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF") if OPENPYXL_AVAILABLE else None
    NORMAL_FONT = Font(size=10) if OPENPYXL_AVAILABLE else None

    # Fills
    HEADER_FILL = PatternFill(
        start_color="1A56DB", end_color="1A56DB", fill_type="solid"
    ) if OPENPYXL_AVAILABLE else None
    TABLE_HEADER_FILL = PatternFill(
        start_color="3B82F6", end_color="3B82F6", fill_type="solid"
    ) if OPENPYXL_AVAILABLE else None
    ALTERNATING_FILL = PatternFill(
        start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"
    ) if OPENPYXL_AVAILABLE else None
    HIGH_SCORE_FILL = PatternFill(
        start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"
    ) if OPENPYXL_AVAILABLE else None
    MEDIUM_SCORE_FILL = PatternFill(
        start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
    ) if OPENPYXL_AVAILABLE else None
    LOW_SCORE_FILL = PatternFill(
        start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"
    ) if OPENPYXL_AVAILABLE else None

    # Borders
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    ) if OPENPYXL_AVAILABLE else None

    # Alignment
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center") if OPENPYXL_AVAILABLE else None
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center") if OPENPYXL_AVAILABLE else None
    WRAP_ALIGN = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    ) if OPENPYXL_AVAILABLE else None


# =============================================================================
# EXCEL GENERATOR CLASS
# =============================================================================


class ExcelGenerator:
    """
    Excel 리포트 생성기.

    ReportData를 Excel 워크북으로 변환한다.

    Attributes:
        _styles: ExcelStyles 인스턴스

    Example:
        >>> generator = ExcelGenerator()
        >>> excel_bytes = generator.generate(report_data)
        >>> generator.save(excel_bytes, "report.xlsx")
    """

    def __init__(self) -> None:
        """
        Excel 생성기 초기화.

        Raises:
            ImportError: openpyxl이 설치되지 않은 경우
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install with: pip install openpyxl"
            )
        self._styles = ExcelStyles()

    def __repr__(self) -> str:
        """String representation for debugging."""
        return "ExcelGenerator()"

    # =========================================================================
    # PUBLIC METHODS
    # =========================================================================

    def generate(self, report: ReportData) -> bytes:
        """
        ReportData를 Excel로 변환한다.

        Args:
            report: ReportData 인스턴스

        Returns:
            Excel 바이트 데이터
        """
        wb = Workbook()

        # Sheet 1: Summary
        self._create_summary_sheet(wb, report)

        # Sheet 2: Business Opportunities
        self._create_opportunities_sheet(wb, report)

        # Sheet 3: Keywords
        self._create_keywords_sheet(wb, report)

        # Sheet 4: Trends
        self._create_trends_sheet(wb, report)

        # Sheet 5: Demands
        self._create_demands_sheet(wb, report)

        # Remove default empty sheet if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.read()

    def save(self, excel_bytes: bytes, filepath: str) -> None:
        """
        Excel을 파일로 저장한다.

        Args:
            excel_bytes: Excel 바이트 데이터
            filepath: 저장할 파일 경로
        """
        with open(filepath, "wb") as f:
            f.write(excel_bytes)

    # =========================================================================
    # PRIVATE METHODS - SHEET CREATION
    # =========================================================================

    def _create_summary_sheet(self, wb: Workbook, report: ReportData) -> None:
        """Summary 시트를 생성한다."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Business Analysis Report: r/{report.subreddit}"
        ws["A1"].font = self._styles.HEADER_FONT
        ws["A1"].fill = self._styles.HEADER_FILL
        ws["A1"].alignment = self._styles.CENTER_ALIGN
        ws.row_dimensions[1].height = 30

        # Metadata
        row = 3
        metadata = [
            ("Generated At", report.generated_at.strftime("%Y-%m-%d %H:%M UTC")),
            ("Analysis Period", report.analysis_period),
            ("Posts Analyzed", f"{report.total_posts_analyzed:,}"),
            ("Total Keywords", str(report.total_keywords)),
            ("Total Insights", str(report.total_insights)),
        ]

        for label, value in metadata:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Executive Summary
        row += 1
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "Executive Summary"
        ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
        row += 1

        ws.merge_cells(f"A{row}:E{row + 3}")
        ws[f"A{row}"] = report.executive_summary
        ws[f"A{row}"].alignment = self._styles.WRAP_ALIGN
        ws.row_dimensions[row].height = 80

        # Market Overview
        row += 5
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "Market Overview"
        ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
        row += 1

        overview = report.market_overview
        overview_items = [
            ("Community", overview.get("community_size", "N/A")),
            ("Activity Level", overview.get("activity_level", "N/A")),
            ("Data Quality", overview.get("data_quality", "N/A")),
            ("Market Maturity", overview.get("market_maturity", "N/A")),
        ]

        for label, value in overview_items:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Key Topics
        if overview.get("key_topics"):
            row += 1
            ws[f"A{row}"] = "Key Topics"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = ", ".join(overview["key_topics"])

        # Recommendations
        row += 3
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "Recommendations"
        ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
        row += 1

        for rec in report.recommendations:
            ws[f"A{row}"] = rec
            ws[f"A{row}"].alignment = self._styles.WRAP_ALIGN
            row += 1

        # Risk Factors
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "Risk Factors"
        ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
        row += 1

        for risk in report.risk_factors:
            ws[f"A{row}"] = f"- {risk}"
            ws[f"A{row}"].alignment = self._styles.WRAP_ALIGN
            row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

    def _create_opportunities_sheet(self, wb: Workbook, report: ReportData) -> None:
        """Business Opportunities 시트를 생성한다."""
        ws = wb.create_sheet("Opportunities")

        # Headers
        headers = [
            "Rank",
            "Title",
            "Category",
            "Score",
            "Market Potential",
            "Risk Level",
            "Target Audience",
            "Description",
            "Key Features",
            "Next Steps",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._styles.TABLE_HEADER_FONT
            cell.fill = self._styles.TABLE_HEADER_FILL
            cell.alignment = self._styles.CENTER_ALIGN
            cell.border = self._styles.THIN_BORDER

        # Data rows
        for row_idx, item in enumerate(report.business_items, 2):
            # Score-based fill
            if item.opportunity_score >= 80:
                fill = self._styles.HIGH_SCORE_FILL
            elif item.opportunity_score >= 60:
                fill = self._styles.MEDIUM_SCORE_FILL
            else:
                fill = self._styles.LOW_SCORE_FILL

            data = [
                item.rank,
                item.title,
                item.category,
                f"{item.opportunity_score:.1f}",
                item.market_potential,
                item.risk_level,
                item.target_audience,
                item.description,
                "\n".join(item.key_features),
                "\n".join(item.next_steps),
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._styles.THIN_BORDER
                cell.alignment = self._styles.WRAP_ALIGN
                if col_idx == 4:  # Score column
                    cell.fill = fill

        # Set column widths
        column_widths = [8, 40, 20, 10, 15, 12, 30, 50, 40, 40]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_keywords_sheet(self, wb: Workbook, report: ReportData) -> None:
        """Keywords 시트를 생성한다."""
        ws = wb.create_sheet("Keywords")

        # Headers
        headers = ["Rank", "Keyword", "Score"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._styles.TABLE_HEADER_FONT
            cell.fill = self._styles.TABLE_HEADER_FILL
            cell.alignment = self._styles.CENTER_ALIGN
            cell.border = self._styles.THIN_BORDER

        # Data rows
        keywords = report.trend_analysis.get("top_keywords", [])
        for row_idx, kw in enumerate(keywords, 2):
            data = [
                row_idx - 1,
                kw.get("keyword", ""),
                f"{kw.get('score', 0):.4f}",
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._styles.THIN_BORDER
                cell.alignment = self._styles.CENTER_ALIGN if col_idx != 2 else self._styles.LEFT_ALIGN

            # Alternating row colors
            if row_idx % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row_idx, column=col).fill = self._styles.ALTERNATING_FILL

        # Set column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add chart if there's data
        if keywords:
            self._add_keyword_chart(ws, len(keywords))

    def _create_trends_sheet(self, wb: Workbook, report: ReportData) -> None:
        """Trends 시트를 생성한다."""
        ws = wb.create_sheet("Trends")

        # Trend Summary
        ws["A1"] = "Trend Summary"
        ws["A1"].font = self._styles.SUBHEADER_FONT
        ws.merge_cells("A2:D4")
        ws["A2"] = report.trend_analysis.get("trend_summary", "No trend data available.")
        ws["A2"].alignment = self._styles.WRAP_ALIGN

        # Statistics
        row = 6
        ws[f"A{row}"] = "Trend Statistics"
        ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
        row += 1

        stats = [
            ("Total Keywords", report.trend_analysis.get("total_keywords", 0)),
            ("Rising Topics", len(report.trend_analysis.get("rising_topics", []))),
        ]

        for label, value in stats:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Rising Topics Table
        row += 2
        ws[f"A{row}"] = "Rising Topics"
        ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
        row += 1

        rising = report.trend_analysis.get("rising_topics", [])
        if rising:
            # Headers
            headers = ["Topic", "Trend"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self._styles.TABLE_HEADER_FONT
                cell.fill = self._styles.TABLE_HEADER_FILL
                cell.border = self._styles.THIN_BORDER
            row += 1

            # Data
            for topic in rising:
                ws.cell(row=row, column=1, value=topic.get("topic", "")).border = self._styles.THIN_BORDER
                ws.cell(row=row, column=2, value=topic.get("trend", "")).border = self._styles.THIN_BORDER
                row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_demands_sheet(self, wb: Workbook, report: ReportData) -> None:
        """Demands 시트를 생성한다."""
        ws = wb.create_sheet("Demands")

        demand_analysis = report.demand_analysis

        # Demand Summary
        ws["A1"] = "Demand Summary"
        ws["A1"].font = self._styles.SUBHEADER_FONT
        ws.merge_cells("A2:D4")
        ws["A2"] = demand_analysis.get("demand_summary", "No demand data available.")
        ws["A2"].alignment = self._styles.WRAP_ALIGN

        # Statistics
        row = 6
        ws[f"A{row}"] = "Demand Statistics"
        ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
        row += 1

        ws[f"A{row}"] = "Total Demands"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = demand_analysis.get("total_demands", 0)
        row += 2

        # Category Distribution
        by_category = demand_analysis.get("by_category", {})
        if by_category:
            ws[f"A{row}"] = "Category Distribution"
            ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
            row += 1

            # Headers
            headers = ["Category", "Count"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self._styles.TABLE_HEADER_FONT
                cell.fill = self._styles.TABLE_HEADER_FILL
                cell.border = self._styles.THIN_BORDER
            row += 1

            # Data
            start_row = row
            for category, count in by_category.items():
                ws.cell(row=row, column=1, value=category).border = self._styles.THIN_BORDER
                ws.cell(row=row, column=2, value=count).border = self._styles.THIN_BORDER
                row += 1

            # Add pie chart for category distribution
            if len(by_category) > 0:
                self._add_category_chart(ws, start_row, row - 1)

        # Top Opportunities
        opportunities = demand_analysis.get("top_opportunities", [])
        if opportunities:
            row += 3
            ws[f"A{row}"] = "Top Demand Opportunities"
            ws[f"A{row}"].font = self._styles.SUBHEADER_FONT
            row += 1

            # Headers
            headers = ["Representative", "Priority Score", "Business Potential"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self._styles.TABLE_HEADER_FONT
                cell.fill = self._styles.TABLE_HEADER_FILL
                cell.border = self._styles.THIN_BORDER
            row += 1

            # Data
            for opp in opportunities[:10]:
                ws.cell(row=row, column=1, value=opp.get("representative", "")[:100]).border = self._styles.THIN_BORDER
                ws.cell(row=row, column=2, value=opp.get("priority_score", 0)).border = self._styles.THIN_BORDER
                ws.cell(row=row, column=3, value=opp.get("business_potential", "")).border = self._styles.THIN_BORDER
                row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20

    # =========================================================================
    # PRIVATE METHODS - CHARTS
    # =========================================================================

    def _add_keyword_chart(self, ws: Worksheet, data_count: int) -> None:
        """키워드 차트를 추가한다."""
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Top Keywords by Score"
        chart.x_axis.title = "Keyword"
        chart.y_axis.title = "Score"

        # Data range (max 10 items for chart clarity)
        max_rows = min(data_count + 1, 11)

        data = Reference(ws, min_col=3, min_row=1, max_row=max_rows)
        categories = Reference(ws, min_col=2, min_row=2, max_row=max_rows)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4

        # Chart position
        ws.add_chart(chart, "E2")

    def _add_category_chart(
        self, ws: Worksheet, start_row: int, end_row: int
    ) -> None:
        """카테고리 파이 차트를 추가한다."""
        chart = PieChart()
        chart.title = "Demand by Category"

        labels = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
        data = Reference(ws, min_col=2, min_row=start_row - 1, max_row=end_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        # Chart position
        ws.add_chart(chart, "D" + str(start_row - 1))


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================


def generate_excel_from_report(report: ReportData) -> bytes:
    """
    ReportData에서 Excel을 생성하는 유틸리티 함수.

    Args:
        report: ReportData 인스턴스

    Returns:
        Excel 바이트 데이터

    Example:
        >>> excel_bytes = generate_excel_from_report(report_data)
    """
    generator = ExcelGenerator()
    return generator.generate(report)
